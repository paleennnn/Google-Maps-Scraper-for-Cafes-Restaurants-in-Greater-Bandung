from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
import os
import random
import logging
import re

# ===== INSTALL OPENPYXL JIKA BELUM ADA =====
try:
    import openpyxl
except ImportError:
    print("[INFO] Installing openpyxl for Excel formatting...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl

# ===== USER-AGENT LIST UNTUK ROTASI =====
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
]

def scrape_gmaps(keyword, limit=50):
    options = Options()
    
    # ===== ANTI-BOT DETECTION MEASURES =====
    user_agent = random.choice(USER_AGENTS)
    options.add_argument(f"user-agent={user_agent}")
    
    # Matikan headless untuk debugging (nyalakan lagi setelah berhasil)
    # options.add_argument("--headless=new")
    
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--start-maximized")
    options.add_argument("--lang=id-ID")  # Set bahasa Indonesia
    
    print(f"[USER-AGENT] {user_agent.split('Chrome/')[1].split()[0] if 'Chrome/' in user_agent else 'Custom'}")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    # Inject stealth JavaScript
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': '''Object.defineProperty(navigator, 'webdriver', {
            get: () => false,
        })'''
    })
    
    print(f"[LOADING] Membuka Google Maps untuk: {keyword}...")
    driver.get(f"https://www.google.com/maps/search/{keyword}")
    
    results = []
    
    try:
        wait = WebDriverWait(driver, 20)
        print("[WAITING] Mencari elemen feed di halaman...")
        
        # Tunggu hingga hasil pencarian muncul
        scrollable_div = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@role="feed"]')))
        print("[FOUND] Feed element ditemukan, mulai scroll...")
        
        # Tunggu sebentar agar konten load
        time.sleep(3)
        
    except Exception as e:
        print(f"[ERROR] Tidak bisa menemukan feed element: {e}")
        driver.quit()
        return results
    
    found = 0
    prev_found = 0
    scroll_attempts = 0
    max_scroll_attempts = 20
    stuck_count = 0
    
    # ===== SCROLL DENGAN DETEKSI STUCK (ANTI-STALE ELEMENT) =====
    merchant_urls = set()
    while len(merchant_urls) < limit and scroll_attempts < max_scroll_attempts:
        try:
            # Scroll ke bawah dengan jarak lebih besar untuk load lebih banyak
            driver.execute_script('arguments[0].scrollTop = arguments[0].scrollHeight', scrollable_div)
            time.sleep(random.uniform(3, 4))
            
            # PERBAIKAN UTAMA: Simpan HREF saja, bukan WebElement
            items = driver.find_elements(By.XPATH, '//a[contains(@href, "/maps/place/")]')
            
            for item in items:
                try:
                    href = item.get_attribute("href")
                    if href and "/maps/place/" in href:
                        # Hapus parameter dinamis untuk deduplikasi yang lebih baik
                        clean_href = href.split("?")[0]
                        merchant_urls.add(clean_href)
                except:
                    continue
            
            found = len(merchant_urls)
            scroll_attempts += 1
            
            print(f"[SCROLL] Attempt {scroll_attempts}: Ditemukan {found} items unik")
            
            # Deteksi jika stuck (tidak ada penambahan item selama 3 scroll)
            if found == prev_found:
                stuck_count += 1
                if stuck_count >= 3:
                    print("[INFO] Sudah mencapai akhir hasil pencarian")
                    break
            else:
                stuck_count = 0
            
            prev_found = found
            
            if found >= limit:
                print(f"[SUCCESS] Target {limit} items tercapai")
                break
                
        except Exception as e:
            print(f"[ERROR] Scroll error: {e}")
            break

    # ===== EKSTRAKSI DATA (ANTI STALE ELEMENT) =====
    print(f"[EXTRACT] Mengekstrak data dari {len(merchant_urls)} merchant unik...")
    
    extracted_count = 0
    failed_count = 0
    
    for idx, url in enumerate(merchant_urls):
        if extracted_count >= limit:
            break
        
        try:
            # Buka URL merchant langsung (SOLUSI CHATGPT - ANTI STALE)
            driver.get(url)
            time.sleep(random.uniform(2.5, 4))
            
            # ===== NAMA MERCHANT =====
            name = "N/A"
            try:
                # Tunggu dan ambil nama dari heading utama
                name = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//h1[contains(@class, "DUwDvf")]')
                    )
                ).text.strip()
            except:
                # Fallback: coba selector alternatif
                try:
                    name = driver.find_element(
                        By.XPATH, '//div[@data-attrid="title"]'
                    ).text.strip()
                except:
                    pass
            
            # Skip jika tidak dapat nama
            if name == "N/A" or name == "":
                print(f"[SKIP] ({idx+1}/{len(merchant_urls)}) Tidak dapat nama merchant")
                continue
            
            # ===== ALAMAT (MAKSIMALKAN CAPTURE) =====
            address = "N/A"
            address_selectors = [
                # Priority 1: Button dengan data-item-id address
                '//button[@data-item-id="address"]',
                '//button[contains(@data-item-id, "address")]',
                
                # Priority 2: Div dengan kelas rogA2c (struktur lama)
                '//div[contains(@class, "rogA2c")]//button[1]',
                '//div[contains(@class, "rogA2c")]//span',
                
                # Priority 3: Div yang berisi kota/provinsi
                '//div[contains(text(), "Bandung") or contains(text(), "Jawa")]',
                '//span[contains(text(), "Jalan") or contains(text(), "Jl.")]',
                
                # Priority 4: Struktur alternatif - cari elemen yang punya format alamat
                '//button[@jsname]//div[contains(@class, "fontBody")]',
                '//div[@data-item-id="address"]',
            ]
            
            for selector in address_selectors:
                try:
                    addr_elem = driver.find_element(By.XPATH, selector)
                    candidate = addr_elem.text.strip()
                    if candidate and len(candidate) > 5:  # Validasi minimal panjang
                        address = candidate
                        break
                except:
                    continue
            
            # ===== TELEPON (MAKSIMALKAN CAPTURE) =====
            phone = "N/A"
            phone_selectors = [
                # Priority 1: Button dengan data-item-id phone
                ('//button[@data-item-id="phone:tel:"]', 'text'),
                ('//button[contains(@data-item-id, "phone:tel")]', 'text'),
                ('//button[contains(@data-item-id, "phone")]', 'text'),
                
                # Priority 2: Link dengan href tel:
                ('//a[contains(@href, "tel:")]', 'href'),
                
                # Priority 3: Div dengan data-item-id phone
                ('//div[@data-item-id="phone:tel:"]', 'text'),
                ('//div[contains(@data-item-id, "phone")]', 'text'),
                
                # Priority 4: Cari teks yang dimulai dengan 0 atau +62
                ('//button[contains(text(), "0") or contains(text(), "+62")]', 'text'),
                
                # Priority 5: Cari dalam span atau div yang berisi angka
                ('//span[contains(text(), "0") and contains(text(), "-")]', 'text'),
                ('//div[contains(text(), "0") and contains(text(), "-")]', 'text'),
            ]
            
            for selector, attr_type in phone_selectors:
                try:
                    elem = driver.find_element(By.XPATH, selector)
                    if attr_type == 'text':
                        candidate = elem.text.strip()
                    else:  # href
                        candidate = elem.get_attribute("href").replace("tel:", "").strip()
                    
                    # Validasi: harus ada minimal angka
                    if candidate and any(c.isdigit() for c in candidate):
                        phone = candidate
                        break
                except:
                    continue
            
            # ===== RATING (MAKSIMALKAN CAPTURE) =====
            rating = "N/A"
            rating_selectors = [
                # Priority 1: Div dengan kelas fontDisplayLarge (format rating utama)
                '//div[contains(@class, "fontDisplayLarge")]',
                
                # Priority 2: Aria-label dengan rating
                '//*[contains(@aria-label, "rating") or contains(@aria-label, "bintang")]',
                
                # Priority 3: Span atau div yang berisi angka desimal (x.x format)
                '//span[contains(text(), ".") and string-length(text()) < 5]',
                '//div[contains(text(), ".") and string-length(text()) < 5]',
                
                # Priority 4: Elemen dengan class yang mengandung "rating"
                '//*[contains(@class, "rating")]',
            ]
            
            for selector in rating_selectors:
                try:
                    rating_elem = driver.find_element(By.XPATH, selector)
                    candidate = rating_elem.text.strip()
                    # Validasi: harus format angka.angka (misal 4.5)
                    if candidate and '.' in candidate:
                        rating = candidate
                        break
                except:
                    continue
            
            # ===== WEBSITE (BONUS) =====
            website = "N/A"
            try:
                website = driver.find_element(
                    By.XPATH, '//a[contains(@href, "http") and not(contains(@href, "google")) and not(contains(@href, "maps"))]'
                ).get_attribute("href").strip()
            except:
                pass
            
            # ===== BACKUP EXTRACTION MENGGUNAKAN PAGE TEXT =====
            # Jika data masih banyak yang N/A, coba extract dari page text langsung
            if phone == "N/A" or address == "N/A" or rating == "N/A":
                try:
                    # Ambil semua text dari page
                    page_text = driver.find_element(By.TAG_NAME, "body").text
                    
                    # Extract telepon dari text jika masih kosong
                    if phone == "N/A":
                        import re
                        # Pattern: angka dengan format telepon Indonesia (0xxx atau +62xxx)
                        phone_pattern = r'(\+62|0)[0-9\s\-\.]{8,}'
                        phone_match = re.search(phone_pattern, page_text)
                        if phone_match:
                            phone = phone_match.group(0).strip()
                    
                    # Extract rating dari text jika masih kosong
                    if rating == "N/A":
                        import re
                        # Pattern: angka desimal (misal 4.5)
                        rating_pattern = r'(\d+[\,\.]\d+)\s*(?:dari|/|out of)?'
                        rating_matches = re.findall(rating_pattern, page_text)
                        if rating_matches:
                            rating = rating_matches[0].replace(',', '.')
                except:
                    pass
            
            results.append({
                "Merchant": name,
                "Alamat": address,
                "Telepon": phone,
                "Rating": rating,
                "Website": website,
                "Sumber": "Google Maps"
            })
            
            extracted_count += 1
            print(f"[EXTRACTED] ({idx+1}/{len(merchant_urls)}) {name}")
            
            # Smart delay untuk menghindari rate limiting
            time.sleep(random.uniform(1, 2))
            
        except Exception as e:
            failed_count += 1
            print(f"[SKIP] ({idx+1}/{len(merchant_urls)}) Error: {str(e)[:60]}")
            
            # Skip URL jika ada error, lanjut ke yang berikutnya
            if failed_count >= 5:
                print("[WARNING] Terlalu banyak error, kemungkinan IP ter-ban. Tunggu sebentar...")
                time.sleep(random.uniform(10, 15))
                failed_count = 0
            
            continue
    
    print(f"[DONE] Total {extracted_count} merchant berhasil diambil ({failed_count} gagal)")

    driver.quit()
    return results

def save_to_checkpoint(new_data, filename="scraped_merchants_bandung.csv"):
    """Simpan data ke CSV dengan append mode untuk auto-checkpoint"""
    if not new_data:
        return
    
    if not os.path.isfile(filename):
        df = pd.DataFrame(new_data)
        df.to_csv(filename, index=False, encoding='utf-8-sig')
    else:
        df = pd.DataFrame(new_data)
        df.to_csv(filename, mode='a', header=False, index=False, encoding='utf-8-sig')
    
    print(f"[CHECKPOINT] {len(new_data)} items disimpan ke {filename}")

# Kota Bandung (30 Kecamatan)
kota_bandung = [
    "Andir", "Antapani", "Arcamanik", "Astanaanyar", "Babakan Ciparay", 
    "Bandung Kidul", "Bandung Kulon", "Bandung Wetan", "Batununggal", "Bojongloa Kaler", 
    "Bojongloa Kidul", "Buahbatu", "Cibeunying Kaler", "Cibeunying Kidul", "Cibiru", 
    "Cicendo", "Cidadap", "Cinambo", "Coblong", "Gedebage", 
    "Kiaracondong", "Lengkong", "Mandalajati", "Panyileukan", "Rancasari", 
    "Regol", "Sukajadi", "Sukasari", "Sumur Bandung", "Ujungberung"
]

# Kabupaten Bandung (31 Kecamatan)
kab_bandung = [
    "Arjasari", "Baleendah", "Banjaran", "Bojongsoang", "Cangkuang", 
    "Cicalengka", "Cikancung", "Cilengkrang", "Cileunyi", "Cimaung", 
    "Cimenyan", "Ciparay", "Ciwidey", "Dayeuhkolot", "Ibun", 
    "Katapang", "Kertasari", "Kutawaringin", "Majalaya", "Margaasih", 
    "Margahayu", "Nagreg", "Pacet", "Pameungpeuk", "Pangalengan", 
    "Rancabali", "Rancaekek", "Solokanjeruk", "Soreang"
]

# Kabupaten Bandung Barat (16 Kecamatan)
kab_bandung_barat = [
    "Batujajar", "Cihampelas", "Cikalongwetan", "Cililin", "Cipatat", 
    "Cipeundeuy", "Cipongkor", "Cisarua", "Gununghalu", "Lembang", 
    "Ngamprah", "Padalarang", "Parongpong", "Rongga", "Sindangkerta", "Saguling"
]

# Kota Cimahi (3 Kecamatan)
kota_cimahi = [
    "Cimahi Selatan", "Cimahi Tengah", "Cimahi Utara"
]

# ===== SETUP LOGGING =====
logging.basicConfig(
    filename='error.log',
    level=logging.ERROR,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# ===== FUNGSI MENU =====
def tampilkan_menu_wilayah():
    """Tampilkan menu pilihan wilayah"""
    print("\n" + "=" * 60)
    print("PILIH WILAYAH YANG AKAN DI-SCRAPE")
    print("=" * 60)
    print("1. Kota Bandung (30 Kecamatan)")
    print("2. Kabupaten Bandung (31 Kecamatan)")
    print("3. Kabupaten Bandung Barat (16 Kecamatan)")
    print("4. Kota Cimahi (3 Kecamatan)")
    print("0. Exit")
    print("=" * 60)
    
    pilihan = input("Masukkan pilihan (0-4): ").strip()
    return pilihan

def tampilkan_menu_kecamatan(nama_wilayah, daftar_kecamatan):
    """Tampilkan menu pilihan kecamatan"""
    print("\n" + "=" * 60)
    print(f"PILIH KECAMATAN DI {nama_wilayah.upper()}")
    print("=" * 60)
    
    for idx, kec in enumerate(daftar_kecamatan, 1):
        print(f"{idx}. {kec}")
    
    print("0. Kembali ke menu wilayah")
    print("=" * 60)
    
    pilihan = input("Masukkan pilihan (0-" + str(len(daftar_kecamatan)) + "): ").strip()
    return pilihan

def jalankan_scrape(nama_wilayah, nama_kecamatan, target_types=["Cafe", "Restoran", "Coffee Shop"]):
    """Jalankan scraping untuk satu kecamatan"""
    
    output_file = f"data_di_{nama_kecamatan.lower().replace(' ', '_')}.csv"
    
    print("\n" + "=" * 60)
    print(f"SCRAPING: {nama_kecamatan} - {nama_wilayah}")
    print("=" * 60)
    
    data_kecamatan = []
    
    # PERBAIKAN: Tambahkan nama wilayah ke query untuk presisi geografis
    for t_type in target_types:
        # Format query dengan nama wilayah untuk menghindari hasil dari kota lain
        if "Kota Bandung" in nama_wilayah or "Kabupaten Bandung" in nama_wilayah or "Bandung Barat" in nama_wilayah:
            query = f"{t_type} di {nama_kecamatan} Bandung"
        elif "Cimahi" in nama_wilayah:
            query = f"{t_type} di {nama_kecamatan} Cimahi"
        else:
            query = f"{t_type} di {nama_kecamatan}"
        
        print(f"\n[SEARCHING] {query}...")
        
        try:
            results = scrape_gmaps(query, limit=30)
            
            if results:
                data_kecamatan.extend(results)
                print(f"[SUCCESS] Ditemukan {len(results)} merchant")
            else:
                print(f"[INFO] Tidak ada hasil untuk {query}")
                
        except Exception as e:
            error_msg = f"Error mencari '{query}': {str(e)}"
            print(f"[ERROR] {error_msg}")
            logger.error(error_msg)
            continue
        
        # Delay antar request
        time.sleep(random.uniform(5, 8))
    
    # ===== FILTER DATA BANDUNG ONLY =====
    if data_kecamatan:
        # Filter hanya data yang alamatnya mengandung kata kunci Bandung/Cimahi
        bandung_keywords = ['Bandung', 'Cimahi', 'Jawa Barat', 'West Java', 'Kota Bdg', 'Kab. Bandung']
        
        filtered_data = []
        excluded_count = 0
        
        for item in data_kecamatan:
            alamat = item.get('Alamat', '').lower()
            # Jika alamat mengandung keyword Bandung ATAU alamat N/A (benefit of doubt)
            if any(keyword.lower() in alamat for keyword in bandung_keywords) or alamat == 'n/a':
                filtered_data.append(item)
            else:
                excluded_count += 1
                print(f"[FILTER] Excluded: {item['Merchant']} (Alamat: {item['Alamat']})")
        
        print(f"\n[FILTER] Total excluded: {excluded_count} items (bukan dari area Bandung)")
        print(f"[FILTER] Total valid: {len(filtered_data)} items")
        
        data_kecamatan = filtered_data
    
    # ===== AUTO-SAVE =====
    if data_kecamatan:
        save_to_checkpoint(data_kecamatan, filename=output_file)
        
        # ===== BERSIHKAN DUPLIKAT =====
        print("\n" + "=" * 60)
        print("MEMBERSIHKAN DATA DUPLIKAT & FORMAT EXCEL")
        print("=" * 60)
        
        try:
            final_df = pd.read_csv(output_file, encoding='utf-8-sig')
            print(f"Total data sebelum deduplikasi: {len(final_df)} baris")
            
            final_df.drop_duplicates(subset=['Merchant', 'Alamat'], keep='first', inplace=True)
            print(f"Total data setelah deduplikasi: {len(final_df)} baris")
            
            # ===== FORMAT EXCEL SEPERTI GAMBAR =====
            # Tambahkan kolom No
            final_df.insert(0, 'No', range(1, len(final_df) + 1))
            
            # Reorder kolom sesuai format yang diinginkan
            column_order = ['No', 'Merchant', 'Alamat', 'Telepon', 'Rating', 'Website', 'Sumber']
            
            # Pastikan semua kolom ada
            for col in column_order:
                if col not in final_df.columns:
                    final_df[col] = 'N/A'
            
            final_df = final_df[column_order]
            
            # Export ke Excel dengan styling
            excel_file = f"data_di_{nama_kecamatan.lower().replace(' ', '_')}.xlsx"
            
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Data')
                
                # Akses worksheet untuk styling
                workbook = writer.book
                worksheet = writer.sheets['Data']
                
                # Styling header
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                
                # Apply styling ke header
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Set column widths
                column_widths = {
                    'A': 5,   # No
                    'B': 35,  # Merchant
                    'C': 60,  # Alamat
                    'D': 18,  # Telepon
                    'E': 10,  # Rating
                    'F': 40,  # Website
                    'G': 15   # Sumber
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Border untuk semua cell
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Apply border dan alignment ke semua data
                for row in worksheet.iter_rows(min_row=1, max_row=len(final_df)+1, 
                                               min_col=1, max_col=len(column_order)):
                    for cell in row:
                        cell.border = thin_border
                        if cell.row > 1:  # Skip header
                            if cell.column == 1:  # No column (center)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif cell.column in [4, 5]:  # Telepon & Rating (center)
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif cell.column == 6:  # Website (hyperlink)
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:  # Text columns (left)
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Freeze header row
                worksheet.freeze_panes = 'A2'
            
            print(f"✓ File final disimpan: {excel_file}")
            print(f"✓ Format: Rapi dengan header berwarna, border, dan auto-width")
            
        except Exception as e:
            error_msg = f"Error saat membersihkan data: {str(e)}"
            print(f"[ERROR] {error_msg}")
            logger.error(error_msg)
    else:
        print("\n[WARNING] Tidak ada data yang berhasil diambil!")
    
    print("\n" + "=" * 60)
    print("SCRAPING SELESAI")
    print("=" * 60)

# ===== MAIN PROGRAM =====
def main():
    """Program utama dengan menu interaktif"""
    
    wilayah_list = {
        "1": ("Kota Bandung", kota_bandung),
        "2": ("Kabupaten Bandung", kab_bandung),
        "3": ("Kabupaten Bandung Barat", kab_bandung_barat),
        "4": ("Kota Cimahi", kota_cimahi),
    }
    
    while True:
        pilihan_wilayah = tampilkan_menu_wilayah()
        
        if pilihan_wilayah == "0":
            print("\nTerima kasih telah menggunakan bot scraper ini!")
            break
        
        if pilihan_wilayah not in wilayah_list:
            print("❌ Pilihan tidak valid!")
            continue
        
        nama_wilayah, daftar_kec = wilayah_list[pilihan_wilayah]
        
        while True:
            pilihan_kec = tampilkan_menu_kecamatan(nama_wilayah, daftar_kec)
            
            if pilihan_kec == "0":
                break
            
            try:
                idx = int(pilihan_kec) - 1
                if 0 <= idx < len(daftar_kec):
                    nama_kecamatan = daftar_kec[idx]
                    jalankan_scrape(nama_wilayah, nama_kecamatan)
                    
                    # Tanya apakah ingin scrape kecamatan lain
                    lanjut = input("\nApakah ingin scrape kecamatan lain? (y/n): ").strip().lower()
                    if lanjut != 'y':
                        break
                else:
                    print("❌ Pilihan tidak valid!")
            except ValueError:
                print("❌ Input harus berupa angka!")

if __name__ == "__main__":
    main()